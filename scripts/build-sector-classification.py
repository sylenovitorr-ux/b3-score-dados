#!/usr/bin/env python3
"""Fetch B3's official economic sector spreadsheet and create a traceable map."""
from __future__ import annotations

import json
import re
import sys
import urllib.request
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "data/sector-classification.json"
URL = "https://sistemaswebb3-listados.b3.com.br/listedCompaniesProxy/CompanyCall/GetDownloadIndustryClassification/eyJsYW5ndWFnZSI6InB0LWJyIiwicGFnZU51bWJlciI6MSwicGFnZVNpemUiOjEwMH0="

def root(code: str) -> str:
    match = re.match(r"^([A-Z]{4})", str(code or "").upper())
    return match.group(1) if match else ""

def build(payload: bytes) -> dict:
    book = load_workbook(BytesIO(payload), data_only=True, read_only=True)
    sheet = book.active
    current = {"sector": None, "subsector": None, "industrySegment": None}
    rows = {}
    for values in sheet.iter_rows(values_only=True):
        cells = [str(value).strip() if value is not None else "" for value in values]
        if len(cells) < 6:
            continue
        # B3 download columns: index, sector, subsector, segment, company, code.
        sector, subsector, segment, name, code = cells[1:6]
        if sector and sector.upper() not in {"SETOR", "SETOR ECONÔMICO", "SECTOR", "ECONOMIC SECTOR"}:
            current["sector"] = sector
        if subsector:
            current["subsector"] = subsector
        if segment:
            current["industrySegment"] = segment
        key = root(code)
        if not key or code.upper() in {"CÓDIGO", "CODIGO", "CODE"}:
            continue
        candidate = {**current, "companyName": name or None, "source": "B3 Classificação setorial", "sourceUrl": URL}
        previous = rows.get(key)
        # Only preserve an unambiguous company-root mapping.
        if previous and (previous["sector"], previous["subsector"], previous["industrySegment"]) != (candidate["sector"], candidate["subsector"], candidate["industrySegment"]):
            rows[key] = {"ambiguous": True, "source": "B3 Classificação setorial", "sourceUrl": URL}
        elif not previous:
            rows[key] = candidate
    return {"generatedAt": datetime.now(UTC).isoformat(), "source": "B3 Classificação setorial", "sourceUrl": URL, "refreshPolicy": "consulta oficial semanal da B3; atualização diária do B3 Score", "groups": rows}

if __name__ == "__main__":
    request = urllib.request.Request(URL, headers={"User-Agent": "B3Score/3.0", "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"})
    with urllib.request.urlopen(request, timeout=60) as response:
        data = response.read()
    result = build(data)
    OUTPUT.write_text(json.dumps(result, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")
    print(f"Wrote {sum(not x.get('ambiguous') for x in result['groups'].values())} official B3 sector mappings")
